import re
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

COLUMN_MAP = {
    "n°": "number",
    "nombre": "name",
    "rut": "rut",
    "edad": "age",
    "cargo": "position",
    "centro de trabajo": "work_center",
    "atención": "attention_type",
    "atencion": "attention_type",
    "tiempo": "time_type",
    "dias perdidos": "lost_days",
    "sexo": "sex",
    "tipo": "incident_type",
    "tipificador": "classifier",
    "parte del cuerpo": "body_part",
    "observación": "observation",
    "observacion": "observation",
    "fecha": "date",
    "año": "year",
    "gasto atención": "attention_cost",
    "gasto atencion": "attention_cost",
    "gasto remedios": "medicine_cost",
    "dia no trabajados": "days_not_worked",
    "gasto dia no trabajado": "cost_per_day_not_worked",
    "gasto total": "total_cost",
    "estado": "status",
    "estado final": "final_status",
    "imagen": "image_url",
}

NUMERIC_INT_FIELDS = {"number", "age", "lost_days", "year"}
NUMERIC_FLOAT_FIELDS = {
    "attention_cost",
    "medicine_cost",
    "days_not_worked",
    "cost_per_day_not_worked",
    "total_cost",
}


def _normalize_header(header: str) -> str:
    if not header:
        return ""
    cleaned = re.sub(r"\s+", " ", str(header).strip().lower())
    return cleaned


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def _parse_int(value) -> int | None:
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _parse_float(value, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _parse_string(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def parse_excel(file_content: bytes, filename: str) -> list[dict]:
    wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)

    target_sheet = "HOJA NUEVA FAYMEX"
    if target_sheet in wb.sheetnames:
        ws = wb[target_sheet]
    else:
        ws = wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    raw_headers = rows[0]
    col_mapping: dict[int, str] = {}
    for idx, header in enumerate(raw_headers):
        if header is None:
            continue
        normalized = _normalize_header(header)
        if normalized in COLUMN_MAP:
            col_mapping[idx] = COLUMN_MAP[normalized]

    records = []
    for row in rows[1:]:
        if not row or all(cell is None for cell in row):
            continue

        record = {}
        for col_idx, field_name in col_mapping.items():
            value = row[col_idx] if col_idx < len(row) else None

            if field_name == "date":
                record[field_name] = _parse_date(value)
            elif field_name in NUMERIC_INT_FIELDS:
                record[field_name] = _parse_int(value)
            elif field_name in NUMERIC_FLOAT_FIELDS:
                record[field_name] = _parse_float(value)
            else:
                record[field_name] = _parse_string(value)

        if not record.get("name") and not record.get("rut") and not record.get("number"):
            continue

        for f in NUMERIC_FLOAT_FIELDS:
            if f not in record:
                record[f] = 0.0
        for f in NUMERIC_INT_FIELDS:
            if f not in record:
                record[f] = None
        if "lost_days" not in record or record["lost_days"] is None:
            record["lost_days"] = 0

        records.append(record)

    return records
